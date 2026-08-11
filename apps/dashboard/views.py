import csv
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from apps.usuarios.models import Usuario
from apps.servicios.models import Servicio, RegistroRSD, RegistroEscombros, RegistroReciclables
from apps.empresas.models import Empresa


def _es_admin(user):
    return user.rol == 'admin' or user.is_staff or user.rol == 'gerencia'


@method_decorator(login_required, name='dispatch')
class DashboardPageView(View):
    def get(self, request):
        if request.user.rol == 'empresa' or request.user.rol == 'cliente':
            return redirect('empresa-portal')
        if request.user.rol in ('recolector', 'operador'):
            return redirect('dashboard-operador')
        return redirect('dashboard-admin')


@method_decorator(login_required, name='dispatch')
class DashboardAdminView(View):
    template_name = 'dashboard/admin.html'

    def get(self, request):
        if not _es_admin(request.user):
            return redirect('dashboard-operador')

        hace_30 = timezone.now() - timedelta(days=30)
        
        # Filtros GET (Fase 8)
        f_empresa  = request.GET.get('empresa')
        f_modulo   = request.GET.get('modulo')
        f_operador = request.GET.get('operador')
        f_estado   = request.GET.get('estado')
        f_inicio   = request.GET.get('inicio')
        f_fin      = request.GET.get('fin')

        # Base queryset
        servicios = Servicio.objects.all()

        if f_empresa:
            servicios = servicios.filter(empresa_id=f_empresa)
        if f_modulo:
            servicios = servicios.filter(modulo=f_modulo)
        if f_operador:
            servicios = servicios.filter(operador_id=f_operador)
        if f_estado:
            servicios = servicios.filter(estado=f_estado)
        if f_inicio:
            servicios = servicios.filter(fecha_retiro_real__date__gte=f_inicio)
        if f_fin:
            servicios = servicios.filter(fecha_retiro_real__date__lte=f_fin)

        validados = servicios.filter(estado='validado')
        pendientes = servicios.filter(estado='pendiente_validacion')
        
        # Totales por módulo
        rsd_total_kg = validados.filter(modulo='rsd', registro_rsd_set__isnull=False).aggregate(total=Sum('registro_rsd_set__cantidad_kg'))['total'] or 0
        escombros_total = validados.filter(modulo='escombros').count()
        reciclables_kg = validados.filter(modulo='reciclables', registro_reciclables_set__isnull=False).aggregate(total=Sum('registro_reciclables_set__cantidad_kg'))['total'] or 0
        
        # Eco-equivalencia global (Solo Reciclables)
        eco_agua = eco_co2 = eco_energia = eco_arboles = 0
        recic_qs = RegistroReciclables.objects.filter(servicio__in=validados)
        for reg in recic_qs:
            eco_agua += reg.eco_agua_L or 0
            eco_co2 += reg.eco_co2_kg or 0
            eco_energia += reg.eco_energia_kwh or 0
            eco_arboles += reg.eco_arboles or 0

        # KPIs mensuales
        lotes_mes = validados.filter(fecha_retiro_real__gte=hace_30).count()
        
        # Top Empresas
        empresas_data = Empresa.objects.annotate(
            total_servicios=Count('servicios', filter=Q(servicios__estado='validado'))
        ).order_by('-total_servicios')[:8]
        
        # Top Operadores
        operadores_data = Usuario.objects.filter(rol='operador').annotate(
            total_rec=Count('servicios_asignados', filter=Q(servicios_asignados__estado='validado'))
        ).order_by('-total_rec')[:10]

        ultimos = servicios.order_by('-id')[:10].select_related('empresa', 'operador')

        # Colección de filtros para plantilla
        all_empresas = Empresa.objects.filter(activa=True).order_by('nombre')
        all_operadores = Usuario.objects.filter(rol='operador', is_active=True).order_by('nombre')

        return render(request, self.template_name, {
            'total_servicios': validados.count(),
            'pendientes_validacion': pendientes.count(),
            'lotes_mes': lotes_mes,
            'rsd_total_kg': rsd_total_kg,
            'escombros_total': escombros_total,
            'reciclables_kg': reciclables_kg,
            
            'eco_agua': eco_agua,
            'eco_co2': eco_co2,
            'eco_energia': eco_energia,
            'eco_arboles': float(eco_arboles),
            
            'empresas': empresas_data,
            'operadores': operadores_data,
            'ultimos': ultimos,

            # Opciones de filtros
            'all_empresas': all_empresas,
            'all_operadores': all_operadores,
            'f_empresa': f_empresa,
            'f_modulo': f_modulo,
            'f_operador': f_operador,
            'f_estado': f_estado,
            'f_inicio': f_inicio,
            'f_fin': f_fin,
        })


@method_decorator(login_required, name='dispatch')
class DashboardOperadorView(View):
    template_name = 'dashboard/operador.html'

    def get(self, request):
        hace_30 = timezone.now() - timedelta(days=30)
        mis_servicios = Servicio.objects.filter(operador=request.user).exclude(estado='cancelado')
        
        pendientes = mis_servicios.filter(estado='pendiente')
        hoy = timezone.now().date()
        programados_hoy = mis_servicios.filter(fecha_programada__date=hoy, estado='pendiente')

        # Últimos retiros gestionados
        historial = mis_servicios.filter(estado__in=['pendiente_validacion', 'observado', 'validado']).order_by('-fecha_retiro_real')[:15]

        # Notificaciones recientes
        notifs_recientes = request.user.notificaciones.filter(leida=False).order_by('-fecha_creacion')[:5]

        return render(request, self.template_name, {
            'mis_servicios': mis_servicios.count(),
            'pendientes': pendientes.count(),
            'programados_hoy': programados_hoy,
            'historial': historial,
            'notifs_recientes': notifs_recientes,
        })


@method_decorator(login_required, name='dispatch')
class DetalleOperadorView(View):
    template_name = 'dashboard/detalle_operador.html'

    def get(self, request, pk):
        if not _es_admin(request.user):
            return redirect('dashboard')
        
        operador = get_object_or_404(Usuario, pk=pk, rol='operador')
        servicios = Servicio.objects.filter(operador=operador).order_by('-id')
        validados = servicios.filter(estado='validado')
        
        hace_30 = timezone.now() - timedelta(days=30)
        servicios_mes = validados.filter(fecha_retiro_real__gte=hace_30)

        return render(request, self.template_name, {
            'operador': operador,
            'servicios': servicios[:20],
            'total_servicios': servicios.count(),
            'validados': validados.count(),
            'servicios_mes': servicios_mes.count()
        })


@method_decorator(login_required, name='dispatch')
class ExportarExcelAPIView(View):
    def get(self, request):
        if not _es_admin(request.user):
            return HttpResponse('Acceso denegado', status=403)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Servicios Redimir"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="006BB8", end_color="006BB8", fill_type="solid")
        
        headers = ['ID', 'Fecha Retiro', 'Empresa', 'Dirección', 'Módulo', 'Cantidad/Unidad', 'Operador', 'Estado', 'Ticket']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        qs = Servicio.objects.filter(estado='validado').select_related('empresa', 'operador')
        
        for s in qs:
            reg = s.get_registro()
            cantidad_str = "N/A"
            ticket_str = ""
            if reg:
                if s.modulo == 'rsd':
                    cantidad_str = f"{reg.cantidad_kg} kg"
                    ticket_str = reg.ticket_externo
                elif s.modulo == 'escombros':
                    cantidad_str = f"{reg.cantidad} {reg.get_unidad_display()}"
                    ticket_str = reg.ticket_externo
                elif s.modulo == 'reciclables':
                    cantidad_str = f"{reg.cantidad_kg} kg"
            
            ws.append([
                s.id,
                s.fecha_retiro_real.strftime('%Y-%m-%d %H:%M') if s.fecha_retiro_real else '—',
                s.empresa.nombre if s.empresa else '—',
                s.direccion,
                s.get_modulo_display(),
                cantidad_str,
                s.operador.nombre_completo if s.operador else '—',
                s.get_estado_display(),
                ticket_str
            ])
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Cierre_Redimir.xlsx"'
        wb.save(response)
        return response

@method_decorator(login_required, name='dispatch')
class CierreMensualView(View):
    def get(self, request):
        if not _es_admin(request.user):
            return redirect('dashboard')
        empresas = Empresa.objects.filter(estado='aprobada', activa=True)
        return render(request, 'dashboard/cierre.html', {'empresas': empresas})


@method_decorator(login_required, name='dispatch')
class ExportarZipAPIView(View):
    """
    Generación de paquete documental en formato .ZIP por Empresa y Período.
    Puntos 3 y 9 del Checklist Maestro Redimir.
    """
    def get(self, request):
        if not _es_admin(request.user):
            return HttpResponse('Acceso denegado', status=403)

        empresa_id = request.GET.get('empresa_id')
        inicio     = request.GET.get('inicio')
        fin        = request.GET.get('fin')

        if not empresa_id or not inicio or not fin:
            return HttpResponse('Parámetros empresa_id, inicio y fin requeridos', status=400)

        empresa = get_object_or_404(Empresa, pk=empresa_id)
        servicios = Servicio.objects.filter(
            empresa=empresa,
            estado='validado',
            fecha_retiro_real__date__range=[inicio, fin]
        )

        import zipfile
        import os
        from io import BytesIO
        from apps.certificados.models import Certificado

        zip_buffer = BytesIO()
        root_dir = f"Redimir_Cierre_{empresa.rut}_{inicio}_al_{fin}"

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 1. Separar servicios por módulo
            servicios_rsd = servicios.filter(modulo='rsd')
            servicios_rescon = servicios.filter(modulo='escombros')
            servicios_recic = servicios.filter(modulo='reciclables')

            # --- MÓDULO RSD ---
            if servicios_rsd.exists():
                # Excel
                wb_rsd = openpyxl.Workbook()
                ws_rsd = wb_rsd.active
                ws_rsd.title = "Tickets RSD"
                ws_rsd.append(['ID Servicio', 'Fecha Retiro', 'Cantidad (kg)', 'Ticket Externo', 'Destino Receptor', 'Operador'])
                for s in servicios_rsd:
                    reg = s.get_registro()
                    ws_rsd.append([
                        s.id,
                        s.fecha_retiro_real.strftime('%Y-%m-%d %H:%M') if s.fecha_retiro_real else '—',
                        getattr(reg, 'cantidad_kg', 0) if reg else 0,
                        getattr(reg, 'ticket_externo', '—') if reg else '—',
                        getattr(reg, 'destino_receptor', '—') if reg else '—',
                        s.operador.nombre_completo if s.operador else '—'
                    ])
                buf_rsd = BytesIO()
                wb_rsd.save(buf_rsd)
                zf.writestr(f"{root_dir}/RSD/Registro_Tickets_RSD.xlsx", buf_rsd.getvalue())

                # Fotos evidencia RSD
                for s in servicios_rsd:
                    reg = s.get_registro()
                    if reg and hasattr(reg, 'fotos'):
                        for foto_obj in reg.fotos.all():
                            if foto_obj.foto and os.path.exists(foto_obj.foto.path):
                                fname = os.path.basename(foto_obj.foto.name)
                                zf.write(foto_obj.foto.path, f"{root_dir}/RSD/Evidencias/Servicio_{s.id}_{fname}")

            # --- MÓDULO RESCON / ESCOMBROS ---
            if servicios_rescon.exists():
                wb_res = openpyxl.Workbook()
                ws_res = wb_res.active
                ws_res.title = "Tickets RESCON"
                ws_res.append(['ID Servicio', 'Fecha Retiro', 'Cantidad', 'Unidad', 'Ticket Externo', 'Destino Receptor', 'Operador'])
                for s in servicios_rescon:
                    reg = s.get_registro()
                    ws_res.append([
                        s.id,
                        s.fecha_retiro_real.strftime('%Y-%m-%d %H:%M') if s.fecha_retiro_real else '—',
                        getattr(reg, 'cantidad', 0) if reg else 0,
                        reg.get_unidad_display() if reg and hasattr(reg, 'get_unidad_display') else 'm3',
                        getattr(reg, 'ticket_externo', '—') if reg else '—',
                        getattr(reg, 'destino_receptor', '—') if reg else '—',
                        s.operador.nombre_completo if s.operador else '—'
                    ])
                buf_res = BytesIO()
                wb_res.save(buf_res)
                zf.writestr(f"{root_dir}/RESCON/Registro_Tickets_RESCON.xlsx", buf_res.getvalue())

                for s in servicios_rescon:
                    reg = s.get_registro()
                    if reg and hasattr(reg, 'fotos'):
                        for foto_obj in reg.fotos.all():
                            if foto_obj.foto and os.path.exists(foto_obj.foto.path):
                                fname = os.path.basename(foto_obj.foto.name)
                                zf.write(foto_obj.foto.path, f"{root_dir}/RESCON/Evidencias/Servicio_{s.id}_{fname}")

            # --- MÓDULO RECICLABLES ---
            if servicios_recic.exists():
                wb_rec = openpyxl.Workbook()
                ws_rec = wb_rec.active
                ws_rec.title = "Reciclables & EcoEquivalencia"
                ws_rec.append(['ID Servicio', 'Fecha Retiro', 'Material', 'Cantidad (kg)', 'Agua (L)', 'CO2 (kg)', 'Energía (kWh)', 'Árboles'])
                for s in servicios_recic:
                    reg = s.get_registro()
                    ws_rec.append([
                        s.id,
                        s.fecha_retiro_real.strftime('%Y-%m-%d %H:%M') if s.fecha_retiro_real else '—',
                        reg.get_material_display() if reg and hasattr(reg, 'get_material_display') else 'Material',
                        getattr(reg, 'cantidad_kg', 0) if reg else 0,
                        getattr(reg, 'eco_agua_L', 0) if reg else 0,
                        getattr(reg, 'eco_co2_kg', 0) if reg else 0,
                        getattr(reg, 'eco_energia_kwh', 0) if reg else 0,
                        getattr(reg, 'eco_arboles', 0) if reg else 0,
                    ])
                buf_rec = BytesIO()
                wb_rec.save(buf_rec)
                zf.writestr(f"{root_dir}/RECICLABLES/Registro_Reciclables.xlsx", buf_rec.getvalue())

                for s in servicios_recic:
                    reg = s.get_registro()
                    if reg and hasattr(reg, 'fotos'):
                        for foto_obj in reg.fotos.all():
                            if foto_obj.foto and os.path.exists(foto_obj.foto.path):
                                fname = os.path.basename(foto_obj.foto.name)
                                zf.write(foto_obj.foto.path, f"{root_dir}/RECICLABLES/Evidencias/Servicio_{s.id}_{fname}")

            # --- CERTIFICADOS PDF POR MÓDULO Y CARPETA CENTRAL ---
            certificados = Certificado.objects.filter(empresa=empresa, periodo_inicio__lte=fin, periodo_fin__gte=inicio)
            for cert in certificados:
                if cert.archivo_pdf and os.path.exists(cert.archivo_pdf.path):
                    fname = os.path.basename(cert.archivo_pdf.path)
                    # Carpeta global Certificados/
                    zf.write(cert.archivo_pdf.path, f"{root_dir}/Certificados/{fname}")
                    
                    # Copia clasificada según los tipos incluidos en el certificado
                    types_str = str(cert.desglose_por_tipo or '').lower()
                    if 'rsd' in types_str or cert.total_rsd_kg > 0:
                        zf.write(cert.archivo_pdf.path, f"{root_dir}/RSD/Certificado_RSD_{cert.codigo_certificado}.pdf")
                    if 'escombros' in types_str or 'rescon' in types_str or cert.total_escombros > 0:
                        zf.write(cert.archivo_pdf.path, f"{root_dir}/RESCON/Certificado_RESCON_{cert.codigo_certificado}.pdf")
                    if 'reciclables' in types_str or 'pet' in types_str or cert.total_reciclables_kg > 0:
                        zf.write(cert.archivo_pdf.path, f"{root_dir}/RECICLABLES/Certificado_Reciclables_{cert.codigo_certificado}.pdf")

            # --- MANIFIESTO DE CIERRE ---
            manifest_content = (
                f"SISTEMA DE GESTIÓN REDIMIR — PAQUETE DOCUMENTAL OFICIAL\n"
                f"==========================================================\n"
                f"Empresa: {empresa.nombre} ({empresa.rut})\n"
                f"Período Certificado: {inicio} al {fin}\n"
                f"Total Servicios Validados: {servicios.count()}\n"
                f"Retiros RSD: {servicios_rsd.count()}\n"
                f"Retiros RESCON: {servicios_rescon.count()}\n"
                f"Retiros Reciclables: {servicios_recic.count()}\n"
                f"Fecha de Generación: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Generado por: {request.user.nombre_completo} ({request.user.rut})\n"
            )
            zf.writestr(f"{root_dir}/MANIFIESTO_DE_CIERRE.txt", manifest_content)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        filename = f"Paquete_Redimir_{empresa.rut}_{inicio}_al_{fin}.zip"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Auditoría de cierre generado
        from apps.usuarios.models import AuditLog
        AuditLog.registrar(
            usuario=request.user,
            accion='emision_doc',
            modelo='Empresa',
            registro_id=empresa.id,
            campo='cierre_mensual',
            valor_nuevo=filename,
            detalles=f"Generado paquete documental ZIP de cierre para {empresa.nombre} del {inicio} al {fin}.",
            ip=request.META.get('REMOTE_ADDR')
        )

        return response

